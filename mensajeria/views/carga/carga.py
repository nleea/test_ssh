from cmath import nan
import re
from types import NoneType
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib.auth.models import User, Group
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required, permission_required
import numpy as np
from mensajeria.models import Archivos, Destinatarios, Personas
from mensajeria.forms import ArchivosForm
import os
from datetime import datetime
from django.conf import settings
import pandas as pd
import json
from django.http import HttpResponse
from openpyxl import Workbook


from rest_framework.generics import CreateAPIView, GenericAPIView
from ...mixins.base import ResponseMixin
from ...serializers.auth.signup_serializers import SignupSerializers
from ...serializers.auth.signin_serializers import SigninSerializers
from rest_framework.response import Response
from rest_framework import status

from mensajeria.tasks import my_task


class Uploaded(CreateAPIView, ResponseMixin):
    serializer_class = SignupSerializers

    def get_validar_numero(self, numero):
        regex = r"\d{10}$"
        return re.match(regex, numero) is not None
    
    def get_binary_search(self, arr, target):
        left, right = 0, len(arr) - 1

        while left <= right:
            mid = (left + right) // 2  

            if arr[mid] == target:
                return mid  
            elif arr[mid] < target:
                left = mid + 1  
            else:
                right = mid - 1  

        return -1
    
    def get_data_null(self, txt):
        if txt is nan:
            return ""
        else:
            return txt
    
    def get_validar_campos(self, matriz):
        return [
            [str(x) if x is not None and not pd.isna(x) else "" for x in row]
            for row in matriz
        ]
    
    def post(self, request, *args, **kwargs):
        
        if "archivo_excel" in request.FILES:
            archivo = request.FILES["archivo_excel"]
            if archivo.name.endswith((".xls", ".xlsx")):
                # try:
                    df = pd.read_excel(archivo, engine="openpyxl")
                    matriz = df.values.tolist()

                    errados             = []
                    validos             = []
                    validos_celular     = []
                    duplicados          = []

                    num_validos      = 0
                    num_duplicados  = 0
                    num_errados     = 0

                    personas_actuales = Personas.objects.all().order_by('telefonowhatsapp')
                    # 


                    validos_actuales = []
                    for persona in personas_actuales:
                        validos_actuales.append(persona.telefonowhatsapp)

                    matriz_data = self.get_validar_campos(matriz)
                    for row in matriz_data:
                        nombre              = row[0]
                        segundo_nombre      = row[1]
                        apellido            = row[2]
                        segundo_apellido    = row[3]
                        celular             = row[4]
                        celular_llamada     = row[5]
                        documento           = row[7]


                        if not isinstance(celular, str):
                            celular         = str(celular)

                        if not isinstance(celular_llamada, str):
                            celular_llamada = str(celular_llamada).rstrip('.0')

                        documento       = str(documento).rstrip('.0')

                        persona_new = {
                                    "nombre"            :   nombre,
                                    "segundo_nombre"    :   segundo_nombre,
                                    "apellido"          :   apellido,
                                    "segundo_apellido"  :   segundo_apellido,
                                    "celular_whatsapp"  :   celular,
                                    "celular_llamada"   :   celular_llamada,
                                    "documento"         :   documento,
                                    "message"           :   ""
                                    }

                        
                        if not self.get_validar_numero(persona_new['celular_whatsapp']):
                            persona_new['message'] = "Numero de whatsapp invalido."
                            errados.append(persona_new)
                            num_errados = num_errados + 1
                        else:        
                            validos_celular.sort()
                            index = self.get_binary_search(validos_celular, celular)

                            if index != -1:
                                persona_new['message'] = "Numero de whatsapp duplicado en el excel."
                                duplicados.append(persona_new)
                                num_duplicados = num_duplicados + 1
                            else: 
                                index2 = self.get_binary_search(validos_actuales, celular)
                                if index2 != -1:
                                    persona_new['message'] = "Numero de whatsapp duplicado en la base de datos."
                                    duplicados.append(persona_new)
                                    num_duplicados = num_duplicados + 1
                                else: 
                                    persona_new['message'] = "Datos correctos."
                                    validos_celular.append(celular)
                                    validos.append(persona_new)
                                    num_validos = num_validos + 1

                    data = {
                        "validos": {"count": num_validos, "data": validos},
                        "errados": {"count": num_errados, "data": errados},
                        "duplicados": {"count": num_duplicados, "data": duplicados},
                    }

                    # result = my_task.delay(duplicados)
                    # result_value = result.get()
                    # print("Tarea")
                    # result = my_task.get()
                    
                    self.data = {"status": status.HTTP_200_OK, "data": data, "state": True}
                # except Exception as e: 
                #     self.data = {"status": status.HTTP_400_BAD_REQUEST, "data": "Error", "state": False}
            
        else:    
            self.data = {"status": status.HTTP_400_BAD_REQUEST, "data": "Error", "state": False}
        return Response(self.response_obj)
    


    
class Save(CreateAPIView, ResponseMixin):
    serializer_class = SignupSerializers

    def post_add_person(self, data, user):

        nueva_persona           = Personas(
            nombre              = data['nombre'],
            segundonombre       = data['segundo_nombre'],
            apellido            = data['apellido'],
            segundoapellido     = data['segundo_apellido'],
            telefonomovil       = data['celular_llamada'],
            telefonowhatsapp    = data['celular_whatsapp'],
            identificacion      = data['documento'],
            created_by          = user
        )

        nueva_persona.save()
        persona_id = nueva_persona.id

        nuevo_registro  = Destinatarios(
            persona_id  = persona_id, 
            created_by  = user, 
            estado_id   = 596
        )
        nuevo_registro.save()

        return True

    def post(self, request, *args, **kwargs):
        personas        = request.data['destinatarios']

        print(personas)
        user            = request.user
        validos         = []
        num_validos     = 0
        invalidos       = []
        num_invalidos   = 0

        
        for persona in personas:

            persona_new = {
                "nombre"            :   persona["nombre"],
                "segundo_nombre"    :   persona["segundo_nombre"],
                "apellido"          :   persona["apellido"],
                "segundo_apellido"  :   persona["segundo_apellido"],
                "celular_whatsapp"  :   persona["celular_whatsapp"],
                "celular_llamada"   :   persona["celular_llamada"],
                "documento"         :   persona["documento"],
                }
            try:
                self.post_add_person(persona_new, user)
                validos.append(persona_new)
                num_validos = num_validos + 1
            except Exception as e:
                invalidos.append(persona_new)
                num_invalidos = num_invalidos + 1


        data = {
                "validos"   : {"count": num_validos, "data": validos},
                "invalidos" : {"count": num_invalidos, "data": invalidos},
                "error"     : num_invalidos
            }



        self.data = {"status": status.HTTP_200_OK, "data": data, "state": True}
        return Response(self.response_obj)

# @login_required(login_url="signin")
# def index(request):
#     #
#     contexto = {
#         "titulo": "Carga",
#         # 'archivos': archivos,
#     }
#     return render(request, "carga/index.html", contexto)


# def uploaded(request):
#     if request.method == "POST" and "archivo_excel" in request.FILES:
#         archivo = request.FILES["archivo_excel"]

#         if archivo.name.endswith((".xls", ".xlsx")):
#             try:
#                 df = pd.read_excel(archivo, engine="openpyxl")
#                 matriz = df.values.tolist()

#                 errados = []
#                 duplicados = []

#                 user = request.user
#                 user = User.objects.get(id=user.id)

#                 for fila in matriz:
#                     nombre = fila[0]
#                     celular = fila[1]

#                     if(nombre !=  ''):
#                         resultado = dividir_nombre(nombre)
#                         if(resultado != 'error'):
#                             if "primer_nombre" in resultado:
#                                 primer_nombre = resultado["primer_nombre"]
#                             else:
#                                 primer_nombre = None

#                             if "segundo_nombre" in resultado:
#                                 segundo_nombre = resultado["segundo_nombre"]
#                             else:
#                                 segundo_nombre = None

#                             if "primer_apellido" in resultado:
#                                 primer_apellido = resultado["primer_apellido"]
#                             else:
#                                 primer_apellido = None

#                             if "segundo_apellido" in resultado:
#                                 segundo_apellido = resultado["segundo_apellido"]
#                             else:
#                                 segundo_apellido = None
#                     else: 
#                         primer_nombre = ""


#                     celular = str(celular)
#                     celular = celular.replace(" ", "")

#                     if (
#                         len(celular) == 10
#                         and celular.startswith("3")
#                         and celular.isdigit()
#                         and primer_nombre != ""
#                         and resultado != 'error'
#                     ):
#                         if not Personas.objects.filter(telefonomovil=celular).exists():

#                             celular_w = '57'+ celular
#                             nueva_persona = Personas(
#                                 nombre=primer_nombre,
#                                 segundonombre=segundo_nombre,
#                                 apellido=primer_apellido,
#                                 segundoapellido=segundo_apellido,
#                                 telefonomovil=celular,
#                                 telefonowhatsapp=celular_w,
#                                 created_by=user
#                             )

#                             nueva_persona.save()
#                             persona_id = nueva_persona.id

#                             nuevo_registro = Destinatarios(
#                                 persona_id=persona_id, created_by=user, estado_id=596
#                             )
#                             nuevo_registro.save()
#                         else:
#                             duplicados.append((nombre, celular))
#                     else:
#                         errados.append((nombre, celular))

#                 archivo_excel = generar_archivo_excel(errados, duplicados)
                
#                 with open(archivo_excel, "rb") as archivo:
#                     response = HttpResponse(
#                         archivo.read(),
#                         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                     )
#                     response[
#                         "Content-Disposition"
#                     ] = f"attachment; filename={archivo_excel}"
#                     return response
                
#             except Exception as e:
                
#                 return JsonResponse(
#                     {"error": "Error al procesar el archivo: " + str(e)}
#                 )

#     return JsonResponse({"error": "No se recibió un archivo válido."})


# def generar_archivo_excel(errados, duplicados):
    
#     workbook = Workbook()
#     hoja_errados = workbook.create_sheet(title="Errados")
#     for registro in errados:
#         hoja_errados.append(registro)

#     hoja_duplicados = workbook.create_sheet(title="Duplicados")
#     for registro in duplicados:
#         hoja_duplicados.append(registro)

#     del workbook["Sheet"]

#     nombre_archivo = "registros.xlsx"
#     workbook.save(nombre_archivo)

#     return nombre_archivo


# def dividir_nombre(nombre):

#     if not nombre:
#         return {"error": 'El parámetro "nombre" es requerido'}

#     try:
#         nombre = nombre.strip()
#     except:
#         return "error"

#     nombre = " ".join(nombre.split())

#     partes = nombre.split()

#     if len(partes) >= 4:
#         primer_nombre = partes[0].capitalize()
#         segundo_nombre = partes[1].capitalize()
#         primer_apellido = partes[2].capitalize()
#         segundo_apellido = partes[3].capitalize()
#     elif len(partes) == 3:
#         primer_nombre = partes[0].capitalize()
#         segundo_nombre = ""
#         primer_apellido = partes[1].capitalize()
#         segundo_apellido = partes[2].capitalize()
#     elif len(partes) == 2:
#         primer_nombre = partes[0].capitalize()
#         segundo_nombre = ""
#         primer_apellido = partes[1].capitalize()
#         segundo_apellido = "" 
#     else:
#         primer_nombre = nombre.capitalize()
#         segundo_nombre = ""
#         primer_apellido = ""  
#         segundo_apellido = ""  


#     resultado = {
#         "primer_nombre": primer_nombre,
#         "segundo_nombre": segundo_nombre,
#         "primer_apellido": primer_apellido,
#         "segundo_apellido": segundo_apellido,
#     }

#     return resultado
